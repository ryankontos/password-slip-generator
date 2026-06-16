#!/usr/bin/env python3
"""Create a sample workbook for testing password-slip-generator."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


OUTPUT = Path(__file__).with_name("sample_passwords.xlsx")

workbook = Workbook()
sheet = workbook.active
sheet.title = "New starters"
headers = [
    "short name",
    "business group",
    "temp password",
    "hiring manager name",
    "start date",
    "office",
]
rows = [
    ["user001", "Group A", "TempPass-001", "Manager 01", "2026-07-01", "Office A"],
    ["user002", "Group B", "TempPass-002", "Manager 02", "2026-07-02", "Office B"],
    ["user003", "Group C", "TempPass-003", "Manager 03", "2026-07-03", "Office C"],
    ["user004", "Group D", "TempPass-004", "Manager 04", "2026-07-04", "Office A"],
    ["user005", "Group A", "TempPass-005", "Manager 05", "2026-07-05", "Office B"],
    ["user006", "Group B", "TempPass-006", "Manager 06", "2026-07-06", "Office C"],
    ["user007", "Group C", "TempPass-007", "Manager 07", "2026-07-07", "Office A"],
    ["user008", "Group D", "TempPass-008", "Manager 08", "2026-07-08", "Office B"],
    ["user009", "Group A", "TempPass-009", "Manager 09", "2026-07-09", "Office C"],
    ["user010", "Group B", "TempPass-010", "Manager 10", "2026-07-10", "Office A"],
    ["user011", "Group C", "TempPass-011", "Manager 11", "2026-07-11", "Office B"],
    ["user012", "Group D", "TempPass-012", "Manager 12", "2026-07-12", "Office C"],
    ["user013", "Group A", "TempPass-013", "Manager 13", "2026-07-13", "Office A"],
    ["user014", "Group B", "TempPass-014", "Manager 14", "2026-07-14", "Office B"],
    ["user015", "Group C", "TempPass-015", "Manager 15", "2026-07-15", "Office C"],
]
sheet.append(headers)
for row in rows:
    sheet.append(row)

for cell in sheet[1]:
    cell.font = Font(color="FFFFFF", bold=True)
    cell.fill = PatternFill("solid", fgColor="1769AA")
for column, width in zip("ABCDEF", [18, 18, 28, 28, 16, 16]):
    sheet.column_dimensions[column].width = width
sheet.freeze_panes = "A2"
sheet.auto_filter.ref = sheet.dimensions

archive = workbook.create_sheet("Previous intake")
archive.append(headers[:4])
archive.append(["sample-user", "Group A", "SamplePass-000", "Manager 00"])

workbook.save(OUTPUT)
print(OUTPUT)
