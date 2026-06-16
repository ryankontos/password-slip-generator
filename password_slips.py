#!/usr/bin/env python3
"""Generate cut-aligned password slip PDFs from the newest Excel file in Downloads."""

from __future__ import annotations

import json
import math
import subprocess
import tkinter as tk
from dataclasses import asdict, dataclass, field
from pathlib import Path
from tkinter import colorchooser, filedialog, messagebox
from typing import Optional

from openpyxl import load_workbook
from reportlab.lib.colors import HexColor
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas


APP_NAME = "password-slip-generator"
SETTINGS_FILE = Path.home() / "Library" / "Application Support" / "Password Slips" / "settings.json"
MM = 72 / 25.4


@dataclass
class Settings:
    workbook: str = ""
    sheet: str = ""
    columns: list[str] = field(default_factory=list)
    output_folder: str = ""

    header_height_mm: float = 20.0
    data_height_mm: float = 20.0
    top_margin_mm: float = 8.0
    bottom_margin_mm: float = 8.0
    side_margin_mm: float = 5.0
    column_gap_mm: float = 2.0
    padding_mm: float = 2.0
    cut_tick_mm: float = 4.0

    header_color: str = "#1769AA"
    header_font_pt: float = 11.0
    data_font_pt: float = 12.0
    minimum_font_pt: float = 4.0

    @property
    def slip_height_mm(self) -> float:
        return self.header_height_mm + self.data_height_mm


def downloads_folder() -> Path:
    return Path.home() / "Downloads"


def newest_downloaded_workbook() -> Optional[Path]:
    folder = downloads_folder()
    if not folder.is_dir():
        return None

    files = [
        file for file in folder.iterdir()
        if file.is_file()
        and file.suffix.lower() in {".xlsx", ".xlsm"}
        and not file.name.startswith((".", "~$"))
    ]
    return max(files, key=lambda file: file.stat().st_mtime) if files else None


def read_saved_settings() -> Settings:
    try:
        data = json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))
        saved = data.get("last_run") or (data.get("recent_runs") or [None])[0]
        settings = Settings(**migrate_saved_settings(saved or {}))
    except (OSError, TypeError, ValueError):
        settings = Settings()

    settings.output_folder = settings.output_folder or str(downloads_folder())

    # Downloads always wins on launch. Column choices still persist by name.
    latest = newest_downloaded_workbook()
    if latest:
        settings.workbook = str(latest)

    return settings


def migrate_saved_settings(data: dict) -> dict:
    renamed = dict(data)
    if "output" in renamed and "output_folder" not in renamed:
        renamed["output_folder"] = str(Path(renamed["output"]).expanduser().parent)
    if "horizontal_padding_mm" in renamed and "padding_mm" not in renamed:
        renamed["padding_mm"] = renamed.pop("horizontal_padding_mm")
    if "header_font_max_pt" in renamed and "header_font_pt" not in renamed:
        renamed["header_font_pt"] = renamed.pop("header_font_max_pt")
    if "data_font_max_pt" in renamed and "data_font_pt" not in renamed:
        renamed["data_font_pt"] = renamed.pop("data_font_max_pt")
    if "font_min_pt" in renamed and "minimum_font_pt" not in renamed:
        renamed["minimum_font_pt"] = renamed.pop("font_min_pt")

    valid_names = set(Settings.__dataclass_fields__)
    return {key: value for key, value in renamed.items() if key in valid_names}


def save_settings(settings: Settings) -> None:
    SETTINGS_FILE.parent.mkdir(parents=True, exist_ok=True)
    SETTINGS_FILE.write_text(json.dumps({"last_run": asdict(settings)}, indent=2), encoding="utf-8")


def save_field_choices(settings: Settings, columns: list[str]) -> None:
    settings.columns = list(columns)
    save_settings(settings)


def workbook_sheet_names(path: str) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return workbook.sheetnames
    finally:
        workbook.close()


def workbook_headers(path: str, sheet_name: str) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        row = next(workbook[sheet_name].iter_rows(min_row=1, max_row=1, values_only=True), ())
        return unique_labels(row)
    finally:
        workbook.close()


def workbook_records(path: str, sheet_name: str, columns: list[str]) -> list[list[str]]:
    if not columns:
        return []

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        headers = unique_labels(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ()))
        indexes = [headers.index(column) for column in columns]
        records = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            values = ["" if index >= len(row) or row[index] is None else str(row[index]) for index in indexes]
            if any(value.strip() for value in values):
                records.append(values)
        return records
    finally:
        workbook.close()


def unique_labels(values) -> list[str]:
    labels = []
    counts: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        label = str(value).strip() if value is not None else f"Column {index}"
        counts[label] = counts.get(label, 0) + 1
        labels.append(label if counts[label] == 1 else f"{label} ({counts[label]})")
    return labels


def slips_per_page(settings: Settings) -> int:
    usable_height = 297 - settings.top_margin_mm - settings.bottom_margin_mm
    if settings.slip_height_mm <= 0 or usable_height <= 0:
        return 0
    return int(usable_height // settings.slip_height_mm)


def page_count(settings: Settings, slip_count: int) -> int:
    per_page = slips_per_page(settings)
    return math.ceil(slip_count / per_page) if per_page else 0


def column_widths(columns: list[str], records: list[list[str]], available_width: float) -> list[float]:
    if not columns:
        return []

    scores = []
    for index, column in enumerate(columns):
        lengths = [len(column)] + [len(row[index]) for row in records if index < len(row)]
        likely = sorted(lengths)[max(0, math.ceil(len(lengths) * 0.85) - 1)]
        scores.append(max(5, min(32, likely + 1.5)))

    minimum = min(available_width * 0.055, available_width / len(columns) * 0.72)
    leftover = max(0, available_width - minimum * len(columns))
    total = sum(scores)
    return [minimum + leftover * score / total for score in scores]


def shrink_to_fit(text: str, font: str, maximum: float, minimum: float, width: float, height: float) -> float:
    size = min(maximum, height / 1.15)
    measured = stringWidth(text, font, size)
    if not text or measured <= width:
        return size
    return max(minimum, min(size, width / max(1, stringWidth(text, font, 1))))


def short_text(text: str, font: str, size: float, width: float) -> str:
    if stringWidth(text, font, size) <= width:
        return text
    suffix = "..."
    if stringWidth(suffix, font, size) > width:
        return ""
    while text and stringWidth(text + suffix, font, size) > width:
        text = text[:-1]
    return text + suffix


def draw_pdf_text(pdf: canvas.Canvas, text: str, font: str, maximum: float, minimum: float,
                  x: float, y: float, width: float, height: float, color) -> None:
    size = shrink_to_fit(text, font, maximum, minimum, width, height)
    pdf.setFillColor(color)
    pdf.setFont(font, size)
    pdf.drawCentredString(x + width / 2, y + (height - size) / 2 + size * 0.18, short_text(text, font, size, width))


def output_path(settings: Settings) -> Path:
    workbook = Path(settings.workbook)
    name = workbook.stem if workbook.name else "password slips"
    return Path(settings.output_folder or downloads_folder()).expanduser() / f"{name} - password slips.pdf"


def make_pdf(settings: Settings) -> tuple[int, int]:
    records = workbook_records(settings.workbook, settings.sheet, settings.columns)
    if not records:
        raise ValueError("No slips to generate.")

    per_page = slips_per_page(settings)
    if per_page < 1:
        raise ValueError("The slip height and margins do not fit on A4.")

    page_width, page_height = A4
    top = settings.top_margin_mm * MM
    side = settings.side_margin_mm * MM
    header_height = settings.header_height_mm * MM
    data_height = settings.data_height_mm * MM
    slip_height = settings.slip_height_mm * MM
    gap = settings.column_gap_mm * MM
    padding = settings.padding_mm * MM
    content_width = page_width - side * 2
    column_area = content_width - gap * (len(settings.columns) - 1)
    widths = column_widths(settings.columns, records, column_area)

    output = output_path(settings)
    output.parent.mkdir(parents=True, exist_ok=True)
    pdf = canvas.Canvas(str(output), pagesize=A4, pageCompression=1)
    pdf.setTitle(APP_NAME)

    pages = page_count(settings, len(records))
    for page in range(pages):
        draw_cut_ticks(pdf, settings, page_width, page_height, per_page)
        page_records = records[page * per_page:(page + 1) * per_page]
        for slot, record in enumerate(page_records):
            draw_slip(pdf, settings, record, widths, side, page_height - top - slot * slip_height, content_width)
        pdf.showPage()

    pdf.save()
    return len(records), pages


def open_print_dialog(pdf: Path) -> bool:
    script = '''
on run argv
    tell application "Preview"
        activate
        open POSIX file (item 1 of argv)
        delay 0.5
        print front document with print dialog
    end tell
end run
'''
    try:
        subprocess.run(["osascript", "-e", script, str(pdf)], check=True)
        return True
    except Exception:
        subprocess.run(["open", str(pdf)], check=False)
        return False


def draw_cut_ticks(pdf: canvas.Canvas, settings: Settings, page_width: float, page_height: float, per_page: int) -> None:
    top = settings.top_margin_mm * MM
    slip_height = settings.slip_height_mm * MM
    tick = settings.cut_tick_mm * MM
    pdf.setStrokeColor(HexColor("#777777"))
    pdf.setLineWidth(0.35)
    for boundary in range(per_page + 1):
        y = page_height - top - boundary * slip_height
        pdf.line(0, y, tick, y)
        pdf.line(page_width - tick, y, page_width, y)


def draw_slip(pdf: canvas.Canvas, settings: Settings, record: list[str], widths: list[float],
              left: float, top: float, content_width: float) -> None:
    header_height = settings.header_height_mm * MM
    data_height = settings.data_height_mm * MM
    gap = settings.column_gap_mm * MM
    padding = settings.padding_mm * MM
    header_y = top - header_height
    data_y = header_y - data_height

    pdf.setFillColor(HexColor(settings.header_color))
    pdf.rect(left, header_y, content_width, header_height, stroke=0, fill=1)

    x = left
    for index, width in enumerate(widths):
        text_width = max(1, width - padding * 2)
        draw_pdf_text(pdf, settings.columns[index], "Helvetica-Bold", settings.header_font_pt,
                      settings.minimum_font_pt, x + padding, header_y, text_width, header_height, HexColor("#FFFFFF"))
        draw_pdf_text(pdf, record[index], "Helvetica", settings.data_font_pt, settings.minimum_font_pt,
                      x + padding, data_y, text_width, data_height, HexColor("#000000"))
        x += width + gap


class LayoutWindow(tk.Toplevel):
    def __init__(self, app: "App") -> None:
        super().__init__(app.root)
        self.app = app
        self.title("Layout and Preview")
        self.geometry("780x600")
        self.transient(app.root)
        self.configure(bg="#F4F6F8")

        self.vars: dict[str, tk.StringVar] = {}
        self.color = tk.StringVar(value=app.settings.header_color)
        self.build()
        self.preview()

    def build(self) -> None:
        body = tk.Frame(self, bg="#F4F6F8", padx=16, pady=16)
        body.pack(fill="both", expand=True)
        body.columnconfigure(0, weight=1)
        body.rowconfigure(3, weight=1)

        tk.Label(
            body, text="Layout and Preview", bg="#F4F6F8", fg="#111827",
            font=("Helvetica", 20, "bold")
        ).grid(row=0, column=0, sticky="w")

        controls = tk.Frame(body, bg="#F4F6F8")
        controls.grid(row=1, column=0, sticky="ew", pady=(10, 10))
        controls.columnconfigure(0, weight=1)
        controls.columnconfigure(1, weight=1)
        controls.columnconfigure(2, weight=1)

        self.group(controls, 0, "Slip", [
            ("Header height", "header_height_mm"),
            ("Data height", "data_height_mm"),
            ("Header font", "header_font_pt"),
            ("Data font", "data_font_pt"),
            ("Minimum font", "minimum_font_pt"),
        ])
        self.group(controls, 1, "Page", [
            ("Top margin", "top_margin_mm"),
            ("Bottom margin", "bottom_margin_mm"),
            ("Side margin", "side_margin_mm"),
            ("Cut ticks", "cut_tick_mm"),
        ])
        self.group(controls, 2, "Columns", [
            ("Column gap", "column_gap_mm"),
            ("Text padding", "padding_mm"),
        ])

        color_row = tk.Frame(body, bg="#F4F6F8")
        color_row.grid(row=2, column=0, sticky="new", pady=(0, 10))
        tk.Label(color_row, text="Header colour", bg="#F4F6F8", fg="#111827").pack(side="left")
        tk.Entry(color_row, textvariable=self.color, width=10, bg="white", fg="#111827").pack(side="left", padx=8)
        tk.Button(color_row, text="Choose...", command=self.choose_color).pack(side="left")
        self.color.trace_add("write", lambda *_: self.preview())

        preview_box = tk.LabelFrame(body, text="Preview", bg="#FFFFFF", fg="#111827", padx=12, pady=12)
        preview_box.grid(row=3, column=0, sticky="nsew")
        preview_box.columnconfigure(0, weight=1)
        preview_box.rowconfigure(0, weight=1)
        self.canvas = tk.Canvas(preview_box, height=170, background="#F1F3F5", highlightthickness=0)
        self.canvas.grid(row=0, column=0, sticky="nsew")
        self.canvas.bind("<Configure>", lambda _: self.preview())

        buttons = tk.Frame(body, bg="#F4F6F8")
        buttons.grid(row=4, column=0, sticky="ew", pady=(12, 0))
        tk.Button(buttons, text="Cancel", command=self.destroy, padx=12, pady=5).pack(side="right")
        tk.Button(
            buttons, text="Apply", command=self.apply, padx=18, pady=5,
            bg="#1769AA", fg="white", activebackground="#0F578F", activeforeground="white"
        ).pack(side="right", padx=(0, 8))

    def group(self, parent: tk.Widget, column: int, title: str, rows: list[tuple[str, str]]) -> None:
        box = tk.LabelFrame(parent, text=title, bg="#FFFFFF", fg="#111827", padx=10, pady=10)
        box.grid(row=0, column=column, sticky="new", padx=(0 if column == 0 else 8, 0))
        for row, (label, name) in enumerate(rows):
            self.vars[name] = tk.StringVar(value=str(getattr(self.app.settings, name)))
            tk.Label(
                box, bg="#FFFFFF", fg="#111827",
                text=f"{label} (mm)" if "font" not in name else f"{label} (pt)"
            ).grid(row=row, column=0, sticky="w", pady=3)
            tk.Entry(
                box, textvariable=self.vars[name], width=8, bg="white", fg="#111827"
            ).grid(row=row, column=1, sticky="e", padx=(8, 0), pady=3)
            self.vars[name].trace_add("write", lambda *_: self.preview())

    def choose_color(self) -> None:
        picked = colorchooser.askcolor(self.color.get(), parent=self, title="Choose header colour")
        if picked[1]:
            self.color.set(picked[1].upper())

    def read(self) -> Optional[Settings]:
        try:
            settings = Settings(**asdict(self.app.settings))
            for name, variable in self.vars.items():
                setattr(settings, name, float(variable.get()))
            settings.header_color = self.color.get().strip()
            HexColor(settings.header_color)
            if any(getattr(settings, name) <= 0 for name in [
                "header_height_mm", "data_height_mm", "header_font_pt", "data_font_pt", "minimum_font_pt"
            ]):
                return None
            if slips_per_page(settings) < 1:
                return None
            return settings
        except (TypeError, ValueError):
            return None

    def preview(self) -> None:
        if not hasattr(self, "canvas"):
            return
        settings = self.read()
        self.canvas.delete("all")
        if settings is None:
            self.canvas.create_text(380, 85, text="Enter valid layout values.", fill="#8A3B3B", font=("Helvetica", 12))
            return
        self.app.draw_preview(self.canvas, settings)

    def apply(self) -> None:
        settings = self.read()
        if settings is None:
            messagebox.showerror(APP_NAME, "Enter valid layout values.", parent=self)
            return
        self.app.settings = settings
        self.app.update_summary()
        self.destroy()


class App:
    def __init__(self, root: tk.Tk, settings: Settings):
        self.root = root
        self.root.title(APP_NAME)
        self.root.configure(bg="#F4F6F8")
        self.settings = settings
        self.headers: list[str] = []
        self.records: list[list[str]] = []

        self.workbook = tk.StringVar(value=settings.workbook)
        self.sheet = tk.StringVar(value=settings.sheet)
        self.output_folder = tk.StringVar(value=settings.output_folder or str(downloads_folder()))
        self.summary = tk.StringVar(value="0 slips | 0 pages")
        self.sheet_menu: Optional[tk.OptionMenu] = None

        self.build()
        self.load_workbook_if_present()

    def build(self) -> None:
        main = tk.Frame(self.root, bg="#F4F6F8", padx=22, pady=12)
        main.pack(fill="both", expand=True)
        main.columnconfigure(0, weight=1)
        main.rowconfigure(1, weight=1)

        tk.Label(
            main, text=APP_NAME, bg="#F4F6F8", fg="#111827",
            font=("Helvetica", 22, "bold")
        ).grid(row=0, column=0, sticky="w", pady=(0, 8))

        content = tk.Frame(main, bg="#F4F6F8")
        content.grid(row=1, column=0, sticky="nsew")
        content.columnconfigure(0, weight=1)
        content.rowconfigure(1, weight=1)

        self.source_section(content)
        self.fields_section(content)
        self.output_section(content)

        footer = tk.Frame(main, bg="#F4F6F8")
        footer.grid(row=2, column=0, sticky="ew", pady=(10, 0))
        tk.Button(footer, text="Layout and Preview...", command=self.open_layout, padx=12, pady=5).pack(side="left")
        tk.Label(
            footer, text="Created by Ryan Kontos", bg="#F4F6F8", fg="#8A8A8A",
            font=("Helvetica", 9)
        ).pack(side="left", padx=(10, 0))
        tk.Button(
            footer, text="Export PDF", command=self.generate, padx=18, pady=7,
            bg="#1769AA", fg="white", activebackground="#0F578F", activeforeground="white"
        ).pack(side="right")
        tk.Button(footer, text="Print...", command=self.print_slips, padx=14, pady=7).pack(side="right", padx=(0, 8))

    def source_section(self, parent: tk.Widget) -> None:
        box = tk.LabelFrame(parent, text="Source", bg="#FFFFFF", fg="#111827", padx=12, pady=12)
        box.grid(row=0, column=0, sticky="ew")
        box.columnconfigure(1, weight=1)
        self.label(box, "Workbook").grid(row=0, column=0, sticky="w", padx=(0, 10), pady=3)
        tk.Entry(box, textvariable=self.workbook, bg="white", fg="#111827").grid(row=0, column=1, sticky="ew", pady=3)
        tk.Button(box, text="Choose...", command=self.choose_workbook).grid(row=0, column=2, padx=(10, 0), pady=3)
        self.label(box, "Sheet").grid(row=1, column=0, sticky="w", padx=(0, 10), pady=3)
        self.sheet_menu = tk.OptionMenu(box, self.sheet, "")
        self.sheet_menu.configure(bg="white", fg="#111827", activebackground="#E8EEF5", highlightthickness=0)
        self.sheet_menu.grid(row=1, column=1, sticky="ew", pady=3)

    def fields_section(self, parent: tk.Widget) -> None:
        box = tk.LabelFrame(parent, text="Fields and order", bg="#FFFFFF", fg="#111827", padx=12, pady=12)
        box.grid(row=1, column=0, sticky="nsew", pady=(8, 0))
        box.columnconfigure(0, weight=1)
        box.columnconfigure(2, weight=1)
        box.rowconfigure(1, weight=1)
        self.label(box, "Available", muted=True).grid(row=0, column=0, sticky="w")
        self.label(box, "Included (left to right)", muted=True).grid(row=0, column=2, sticky="w")

        self.available = self.listbox(box, 1, 0)
        self.available.bind("<Double-Button-1>", lambda _: self.add_fields())

        buttons = tk.Frame(box, bg="#FFFFFF")
        buttons.grid(row=1, column=1, padx=12)
        tk.Button(buttons, text="Add  >", width=9, command=self.add_fields).pack(pady=5)
        tk.Button(buttons, text="<  Remove", width=9, command=self.remove_fields).pack(pady=5)

        included_frame = tk.Frame(box, bg="#FFFFFF")
        included_frame.grid(row=1, column=2, sticky="nsew", pady=(5, 0))
        included_frame.columnconfigure(0, weight=1)
        included_frame.rowconfigure(0, weight=1)
        self.included = tk.Listbox(
            included_frame, selectmode="extended", exportselection=False, height=6,
            bg="white", fg="#111827", selectbackground="#1769AA", selectforeground="white"
        )
        self.included.grid(row=0, column=0, sticky="nsew")
        scroll = tk.Scrollbar(included_frame, command=self.included.yview)
        scroll.grid(row=0, column=1, sticky="ns")
        self.included.configure(yscrollcommand=scroll.set)
        self.included.bind("<Double-Button-1>", lambda _: self.remove_fields())

        move = tk.Frame(included_frame, bg="#FFFFFF")
        move.grid(row=0, column=2, sticky="n", padx=(8, 0))
        tk.Button(move, text="Move up", width=9, command=lambda: self.move_field(-1)).pack(pady=(0, 5))
        tk.Button(move, text="Move down", width=9, command=lambda: self.move_field(1)).pack(pady=5)

    def output_section(self, parent: tk.Widget) -> None:
        box = tk.LabelFrame(parent, text="Output", bg="#FFFFFF", fg="#111827", padx=12, pady=12)
        box.grid(row=2, column=0, sticky="ew", pady=(8, 0))
        box.columnconfigure(1, weight=1)
        tk.Label(
            box, textvariable=self.summary, bg="#FFFFFF", fg="#111827",
            font=("Helvetica", 15, "bold")
        ).grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 8))
        self.label(box, "Folder").grid(row=1, column=0, sticky="w", padx=(0, 10))
        tk.Entry(box, textvariable=self.output_folder, bg="white", fg="#111827").grid(row=1, column=1, sticky="ew")
        tk.Button(box, text="Choose...", command=self.choose_output_folder).grid(row=1, column=2, padx=(10, 0))

    @staticmethod
    def label(parent: tk.Widget, text: str, muted: bool = False) -> tk.Label:
        return tk.Label(parent, text=text, bg="#FFFFFF", fg="#8A8A8A" if muted else "#111827")

    @staticmethod
    def listbox(parent: tk.Widget, row: int, column: int) -> tk.Listbox:
        frame = tk.Frame(parent, bg="#FFFFFF")
        frame.grid(row=row, column=column, sticky="nsew", pady=(5, 0))
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(0, weight=1)
        box = tk.Listbox(
            frame, selectmode="extended", exportselection=False, height=6,
            bg="white", fg="#111827", selectbackground="#1769AA", selectforeground="white"
        )
        box.grid(row=0, column=0, sticky="nsew")
        scroll = tk.Scrollbar(frame, command=box.yview)
        scroll.grid(row=0, column=1, sticky="ns")
        box.configure(yscrollcommand=scroll.set)
        return box

    def load_workbook_if_present(self) -> None:
        if Path(self.workbook.get()).is_file():
            self.load_sheets()
        else:
            self.update_summary()

    def choose_workbook(self) -> None:
        path = filedialog.askopenfilename(
            title="Choose Excel workbook",
            initialdir=str(downloads_folder()),
            filetypes=[("Excel workbooks", "*.xlsx *.xlsm"), ("All files", "*")],
        )
        if path:
            self.workbook.set(path)
            self.load_sheets()

    def load_sheets(self) -> None:
        try:
            sheets = workbook_sheet_names(self.workbook.get())
            self.update_sheet_menu(sheets)
            if self.sheet.get() not in sheets:
                self.sheet.set(sheets[0])
            self.load_columns()
        except Exception as error:
            messagebox.showerror(APP_NAME, f"Could not open workbook:\n{error}")

    def update_sheet_menu(self, sheets: list[str]) -> None:
        if self.sheet_menu is None:
            return
        menu = self.sheet_menu["menu"]
        menu.delete(0, "end")
        for sheet in sheets:
            menu.add_command(label=sheet, command=lambda value=sheet: self.choose_sheet(value))

    def choose_sheet(self, sheet: str) -> None:
        self.sheet.set(sheet)
        self.load_columns()

    def load_columns(self) -> None:
        try:
            self.headers = workbook_headers(self.workbook.get(), self.sheet.get())
            included = [name for name in self.settings.columns if name in self.headers] or list(self.headers)
            self.fill(self.included, included)
            self.fill(self.available, [name for name in self.headers if name not in included])
            self.refresh_records()
        except Exception as error:
            messagebox.showerror(APP_NAME, f"Could not read sheet:\n{error}")

    @staticmethod
    def fill(box: tk.Listbox, values: list[str]) -> None:
        box.delete(0, "end")
        for value in values:
            box.insert("end", value)

    def included_columns(self) -> list[str]:
        return list(self.included.get(0, "end"))

    def add_fields(self) -> None:
        selected = [self.available.get(index) for index in self.available.curselection()]
        if not selected:
            return
        self.fill(self.included, self.included_columns() + selected)
        self.fill(self.available, [value for value in self.available.get(0, "end") if value not in selected])
        self.refresh_records()
        self.remember_fields()

    def remove_fields(self) -> None:
        selected = [self.included.get(index) for index in self.included.curselection()]
        if not selected:
            return
        included = [value for value in self.included_columns() if value not in selected]
        available = [name for name in self.headers if name in set(self.available.get(0, "end")) | set(selected)]
        self.fill(self.included, included)
        self.fill(self.available, available)
        self.refresh_records()
        self.remember_fields()

    def move_field(self, direction: int) -> None:
        selected = list(self.included.curselection())
        values = self.included_columns()
        if not selected or (direction < 0 and selected[0] == 0) or (direction > 0 and selected[-1] == len(values) - 1):
            return
        indexes = selected if direction < 0 else reversed(selected)
        for index in indexes:
            values[index], values[index + direction] = values[index + direction], values[index]
        self.fill(self.included, values)
        for index in [index + direction for index in selected]:
            self.included.selection_set(index)
        self.refresh_records()
        self.remember_fields()

    def remember_fields(self) -> None:
        self.settings.workbook = self.workbook.get()
        self.settings.sheet = self.sheet.get()
        self.settings.output_folder = self.output_folder.get()
        save_field_choices(self.settings, self.included_columns())

    def refresh_records(self) -> None:
        self.records = workbook_records(self.workbook.get(), self.sheet.get(), self.included_columns())
        self.update_summary()

    def update_summary(self) -> None:
        count = len(self.records)
        per_page = slips_per_page(self.settings)
        pages = page_count(self.settings, count)
        self.summary.set(f"{count} slips  |  {pages} pages  |  {per_page} slips per page" if per_page else "Layout does not fit A4")

    def open_layout(self) -> None:
        for child in self.root.winfo_children():
            if isinstance(child, LayoutWindow):
                child.lift()
                return
        LayoutWindow(self)

    def draw_preview(self, drawing: tk.Canvas, settings: Settings) -> None:
        drawing.delete("all")
        columns = self.included_columns()
        if not columns:
            drawing.create_text(390, 90, text="Add fields to preview.", fill="#666666", font=("Helvetica", 12))
            return

        width = max(500, drawing.winfo_width())
        height = max(150, drawing.winfo_height())
        left, right = 18, width - 18
        top = 16
        slip_height = height - 32
        header_height = slip_height * settings.header_height_mm / settings.slip_height_mm
        gap = max(2, settings.column_gap_mm * 4)
        records = self.records[:1] or [["Sample value"] * len(columns)]
        widths = column_widths(columns, records, right - left - gap * (len(columns) - 1))

        drawing.create_rectangle(left, top, right, top + header_height, fill=settings.header_color, outline="")
        drawing.create_rectangle(left, top + header_height, right, top + slip_height, fill="#FFFFFF", outline="")

        x = left
        for index, column_width in enumerate(widths):
            value = records[0][index] if index < len(records[0]) else ""
            self.canvas_text(drawing, columns[index], x, top, column_width, header_height, "#FFFFFF", True, settings.header_font_pt)
            self.canvas_text(drawing, value, x, top + header_height, column_width, slip_height - header_height, "#111111", False, settings.data_font_pt)
            x += column_width + gap

    @staticmethod
    def canvas_text(drawing: tk.Canvas, text: str, x: float, y: float, width: float, height: float,
                    color: str, bold: bool, max_size: float) -> None:
        size = min(max_size, max(6, height * 0.28))
        while size > 5 and len(text) * size * 0.56 > max(1, width - 14):
            size -= 1
        max_chars = max(1, int((width - 14) / max(1, size * 0.56)))
        shown = text if len(text) <= max_chars else text[:max(1, max_chars - 3)] + "..."
        drawing.create_text(x + width / 2, y + height / 2, text=shown, fill=color,
                            font=("Helvetica", int(size), "bold" if bold else "normal"))

    def choose_output_folder(self) -> None:
        folder = filedialog.askdirectory(title="Choose output folder", initialdir=self.output_folder.get())
        if folder:
            self.output_folder.set(folder)
            self.remember_fields()

    def current_settings(self) -> Settings:
        settings = Settings(**asdict(self.settings))
        settings.workbook = self.workbook.get()
        settings.sheet = self.sheet.get()
        settings.columns = self.included_columns()
        settings.output_folder = self.output_folder.get()
        return settings

    def generate(self) -> None:
        try:
            settings = self.current_settings()
            count, pages = self.generate_pdf(settings)
            messagebox.showinfo(APP_NAME, f"Created {count} slips across {pages} page(s).\n\n{output_path(settings)}")
        except Exception as error:
            messagebox.showerror(APP_NAME, str(error))

    def print_slips(self) -> None:
        try:
            settings = self.current_settings()
            count, pages = self.generate_pdf(settings)
            if not open_print_dialog(output_path(settings)):
                messagebox.showinfo(
                    APP_NAME,
                    f"Created {count} slips across {pages} page(s), then opened the PDF.\n\n"
                    "Use File > Print if the print dialog did not appear."
                )
        except Exception as error:
            messagebox.showerror(APP_NAME, str(error))

    def generate_pdf(self, settings: Settings) -> tuple[int, int]:
        if not Path(settings.workbook).is_file():
            raise ValueError("Choose an Excel workbook.")
        if not settings.columns:
            raise ValueError("Add at least one field.")
        count, pages = make_pdf(settings)
        save_settings(settings)
        self.settings = settings
        return count, pages


def main() -> None:
    root = tk.Tk()
    root.title(APP_NAME)
    root.geometry("940x610")
    root.minsize(800, 540)
    App(root, read_saved_settings())
    root.update_idletasks()
    root.lift()
    root.after(100, root.focus_force)
    root.mainloop()


if __name__ == "__main__":
    main()
